from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from mdconvert_app.assets import AssetManager
from mdconvert_app.markdown_utils import markdown_table, normalize_whitespace
from mdconvert_app.models import ConversionResult


def convert_xlsx(source: Path, destination: Path) -> ConversionResult:
    workbook = load_workbook(source, data_only=True)
    assets = AssetManager(destination)
    lines: list[str] = [f"# {source.stem}", ""]

    for sheet in workbook.worksheets:
        lines.append(f"## {sheet.title}")
        lines.append("")

        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                rows.append([normalize_whitespace("" if cell is None else str(cell)) for cell in row])
        if rows:
            lines.append(markdown_table(rows))
            lines.append("")

        for index, image in enumerate(getattr(sheet, "_images", []), start=1):
            raw = image._data()
            detected = _detect_image_extension(raw)
            relative = assets.save_bytes(raw, f".{detected}", f"{sheet.title}-image-{index}")
            lines.append(f"![{sheet.title} image {index}]({relative})")
            lines.append("")

    markdown = "\n".join(_trim(lines)).strip() + "\n"
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(markdown, encoding="utf-8")
    return ConversionResult(source=source, destination=destination, markdown=markdown)


def _trim(lines: list[str]) -> list[str]:
    trimmed = list(lines)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def _detect_image_extension(raw: bytes) -> str:
    if raw.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if raw.startswith(b"\xff\xd8\xff"):
        return "jpg"
    if raw.startswith((b"GIF87a", b"GIF89a")):
        return "gif"
    if raw.startswith(b"BM"):
        return "bmp"
    if raw.startswith(b"RIFF") and raw[8:12] == b"WEBP":
        return "webp"
    if raw.startswith(b"II*\x00") or raw.startswith(b"MM\x00*"):
        return "tiff"
    return "png"
